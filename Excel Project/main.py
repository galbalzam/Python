import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
workbook = openpyxl.Workbook()
sheet = workbook.active
path = "Grades_table.xlsx"
file_object = openpyxl.load_workbook(path)
sheet_obj = file_object.active


def update():
    #   Define headers
    header_name = sheet.cell(row=1, column=1)
    header_name.value = "Name"
    header_name.font = Font(size="14", bold=True)

    header_grade = sheet.cell(row=1, column=2)
    header_grade.value = "Grade"
    header_grade.font = Font(size="14", bold=True)

    #   Insert names and grades
    cnt = 2
    while True:
        name = input("Enter Student Name ('finish' to end): ")
        if name == "finish":
            break
        cell_name = sheet.cell(row=cnt, column=1)
        cell_name.value = name
        grade = int(input("Enter grade: "))
        cell_grade = sheet.cell(row=cnt, column=2)
        cell_grade.value = grade
        cnt += 1

    #   Save file
    try:
        workbook.save(filename="Grades_table.xlsx")
    except:
        print("The File Is Open")
    else:
        print("File Updated - Grades_table.xlsx")


def max_garde():
    path = "Grades_table.xlsx"
    file_object = openpyxl.load_workbook(path)
    sheet_obj = file_object.active
    max_grade = sheet_obj.cell(row=2, column=2).value
    for i in range(2, sheet_obj.max_row + 1):
        if max_grade < sheet_obj.cell(row=i, column=2).value:
            max_grade = sheet_obj.cell(row=i, column=2).value
    return max_grade


def min_grade():
    path = "Grades_table.xlsx"
    file_object = openpyxl.load_workbook(path)
    sheet_obj = file_object.active
    minimum_grade = sheet_obj.cell(row=2, column=2).value
    for i in range(2, sheet_obj.max_row + 1):
        if minimum_grade > sheet_obj.cell(row=i, column=2).value:
            minimum_grade = sheet_obj.cell(row=i, column=2).value
    return minimum_grade


def avg_garde():
    path = "Grades_table.xlsx"
    file_object = openpyxl.load_workbook(path)
    sheet_obj = file_object.active
    sum_grades = 0
    # calculate sum of grades
    for i in range(2, sheet_obj.max_row + 1):
        sum_grades += sheet_obj.cell(row=i, column=2).value
    # calculate average
    avg = sum_grades / (sheet_obj.max_row - 1)
    return avg

def add_barchar_V2():
    path = "Grades_table.xlsx"
    file_object = openpyxl.load_workbook(path)
    sheet_obj = file_object.active
    barchar_workbook = openpyxl.Workbook()
    barchar_sheet = barchar_workbook.active
    for i in range(2, sheet_obj.max_row + 1):
        grade = sheet_obj.cell(row=i, column=2)
        print(grade.value)
        barchar_sheet.append([grade.value])

    values = Reference(sheet_obj, min_col=1, min_row=1, max_col=15, max_row=sheet_obj.max_row)
    print("sheet_obj max:", sheet_obj.max_row)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    # chart.x_axis.title = " X_AXIS "
    chart.y_axis.title = "Grade"
    barchar_sheet.add_chart(chart, "E3")
    try:
        barchar_workbook.save(filename="barchar_v2.xlsx")
    except:
        print("File is open")
    else:
        print("File Updated - barchar_v2.xlsx ")




def print_all_grades():
    path = "Grades_table.xlsx"
    file_object = openpyxl.load_workbook(path)
    sheet_obj = file_object.active
    for i in range(2, sheet_obj.max_row + 1):
        print(sheet_obj.cell(row=i, column=2).value)


#   Define client menu
while True:
    print("----------MENU----------")
    print("Choose Option - Press 0 to finish")
    print("1. Update Students And Grades Table")
    print("2. Print Best Grade")
    print("3. Print Smallest Grade")
    print("4. Print Grades Average")
    print("5. Graph")
    print("6. Print All The Grades")

    print("0. Finish")

    # Get input from client
    option = int(input("Your Option: "))
    if option == 0:
        break
    elif option == 1:
        update()
    elif option == 2:
        print("The best grade is:", max_garde())
    elif option == 3:
        print("The smallest grade is: ", min_grade())
    elif option == 4:
        print("The avg garde is: ", avg_garde())
    elif option == 5:
        add_barchar_V2()
    elif option == 6:
        print("The grades are: ")
        print_all_grades()
